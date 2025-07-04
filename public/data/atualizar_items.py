import csv
import json
import os

try:
    from openpyxl import load_workbook
except ImportError:  # openpyxl é opcional
    load_workbook = None

# Caminhos dos arquivos
DIR = os.path.dirname(__file__)
CAMINHO_CSV = os.path.join(DIR, 'CSV-itens.csv')
CAMINHO_XLSX = os.path.join(DIR, 'CSV-itens.xlsx')
CAMINHO_JSON = os.path.join(DIR, 'items.json')  # arquivo JSON gerado


def carregar_planilha_para_lista():
    """Carrega CSV ou XLSX para uma lista de dicionários"""
    if os.path.exists(CAMINHO_XLSX) and load_workbook is not None:
        wb = load_workbook(CAMINHO_XLSX)
        sheet = wb.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        linhas = (
            dict(zip(headers, [cell.value for cell in row]))
            for row in sheet.iter_rows(min_row=2)
        )
    else:
        with open(CAMINHO_CSV, mode='r', encoding='utf-8') as f:
            linhas = list(csv.DictReader(f))

    return [
        {
            "nome": linha.get("nome", ""),
            "width": int(linha.get("width", 0)),
            "height": int(linha.get("height", 0)),
            "color": linha.get("color", ""),
            "img": linha.get("img", ""),
            "maxEstresse": int(linha.get("maxEstresse", 3)),
            "tipo": linha.get("tipo", ""),
            "slot": linha.get("slot", "")
        }
        for linha in linhas
    ]


def atualizar_arquivo_json(caminho, novos_itens):
    # Se o arquivo existir, tenta carregar o conteúdo anterior
    if os.path.exists(caminho):
        with open(caminho, mode='r', encoding='utf-8') as f:
            try:
                conteudo = json.load(f)
            except json.JSONDecodeError:
                conteudo = []
    else:
        conteudo = []

    # Substituir todos os itens (ou você pode fazer merge se quiser)
    conteudo = novos_itens

    with open(caminho, mode='w', encoding='utf-8') as f:
        json.dump(conteudo, f, indent=2, ensure_ascii=False)
    print(f"Arquivo '{caminho}' atualizado com {len(novos_itens)} itens.")


# Execução
itens = carregar_planilha_para_lista()
atualizar_arquivo_json(CAMINHO_JSON, itens)
